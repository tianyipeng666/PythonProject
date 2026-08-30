from __future__ import annotations

import unittest
from datetime import datetime

import pandas as pd
from openpyxl import Workbook

from think.etf.models import MarketMove, MetricReference, ValuationResult
from think.etf.analyzer import analyze_index
from think.etf.data import LIXINGER_SOURCE, OfficialSnapshot
from think.etf.config import INDEX_SPECS
from think.etf.data import DataProvider
from think.etf.database import parse_mysql_url
from think.etf.strategy import contribution_multiplier
from think.etf.valuation import (
    absolute_yield_score,
    classify_valuation,
    combined_valuation_score,
    metric_reference,
)
from think.etf.report import _build_detail


class ValuationTest(unittest.TestCase):
    def test_metric_reference_and_percentile(self):
        frame = pd.DataFrame(
            {
                "date": pd.date_range("2020-01-01", periods=100, freq="30D"),
                "value": list(range(1, 101)),
            }
        )
        reference = metric_reference(frame, history_years=20)
        self.assertEqual(reference.current, 100)
        self.assertEqual(reference.percentile, 100)
        self.assertAlmostEqual(reference.q20, 20.8)
        self.assertEqual(reference.sample_count, 100)

    def test_classification_boundaries(self):
        self.assertEqual(classify_valuation(20), "明显低估")
        self.assertEqual(classify_valuation(35), "偏低")
        self.assertEqual(classify_valuation(65), "合理")
        self.assertEqual(classify_valuation(80), "偏高")
        self.assertEqual(classify_valuation(81), "明显高估")

    def test_large_yield_spread_offsets_high_relative_percentile(self):
        self.assertEqual(absolute_yield_score(9.0), 15.0)
        score = combined_valuation_score(relative_percentile=95.0, earnings_yield_spread=9.0)
        self.assertLess(score, 80.0)


class StrategyTest(unittest.TestCase):
    @staticmethod
    def move(return_5d: float, drawdown_60d: float) -> MarketMove:
        return MarketMove(
            latest_date=datetime.now().date().isoformat(),
            close=100.0,
            return_1d=-0.01,
            return_5d=return_5d,
            return_20d=-0.10,
            drawdown_60d=drawdown_60d,
            drawdown_250d=-0.20,
            volatility_20d=0.30,
        )

    def test_low_valuation_gets_dip_bonus(self):
        base, bonus, _ = contribution_multiplier("偏低", self.move(-0.09, -0.13))
        self.assertEqual(base, 1.20)
        self.assertEqual(bonus, 0.40)

    def test_high_valuation_does_not_chase_dip(self):
        base, bonus, reason = contribution_multiplier("偏高", self.move(-0.15, -0.20))
        self.assertEqual(base, 0.75)
        self.assertEqual(bonus, 0.0)
        self.assertIn("估值偏高", reason)

    def test_fair_valuation_caps_bonus(self):
        base, bonus, _ = contribution_multiplier("合理", self.move(-0.15, -0.20))
        self.assertEqual(base, 1.0)
        self.assertEqual(bonus, 0.25)


class ExactDataTest(unittest.TestCase):
    def test_mysql_exact_history_replaces_proxy(self):
        frame = pd.DataFrame(
            {
                "date": pd.date_range("2024-01-01", periods=24, freq="MS"),
                "pe": [30 + i / 10 for i in range(24)],
                "pb": [4 + i / 100 for i in range(24)],
                "dividend_yield": [0.8] * 24,
                "quality": ["exact"] * 24,
            }
        )

        class FakeRepository:
            def best_exact_source(self, code, min_rows, min_history_days):
                return "wind"

            def load_valuations(self, code, source):
                return frame

            def latest_official_valuation(self, code):
                row = frame.iloc[-1]
                return {
                    "date": row["date"].date().isoformat(),
                    "pe": row["pe"], "pb": row["pb"],
                    "dividend_yield": row["dividend_yield"], "source": "wind",
                }

        provider = DataProvider(repository=FakeRepository(), offline=True)
        spec = INDEX_SPECS["399006"]
        pe = provider.pe_history(spec)
        pb = provider.pb_history(spec)
        snapshot = provider.official_snapshot(spec)
        self.assertEqual(len(pe), 24)
        self.assertIsNotNone(pb)
        self.assertTrue(provider.using_exact("399006"))
        self.assertAlmostEqual(snapshot.pe, 32.3)

    def test_lixinger_frame_maps_metrics_and_converts_dividend_ratio(self):
        frame = DataProvider._lixinger_frame(
            [
                {
                    "date": "2026-08-27T00:00:00+08:00",
                    "stockCode": "000300",
                    "pe_ttm.mcw": 14.12,
                    "pb.mcw": 1.46,
                    "dyr.mcw": 0.027,
                }
            ]
        )
        self.assertEqual(len(frame), 1)
        self.assertAlmostEqual(frame.iloc[0]["pe"], 14.12)
        self.assertAlmostEqual(frame.iloc[0]["pb"], 1.46)
        self.assertAlmostEqual(frame.iloc[0]["dividend_yield"], 2.7)

    def test_lixinger_memory_history_is_exact_without_database(self):
        provider = DataProvider(repository=None, offline=True)
        frame = pd.DataFrame(
            {
                "date": pd.date_range("2020-01-01", periods=60, freq="MS"),
                "pe": [12.0] * 60,
                "pb": [1.2] * 60,
                "dividend_yield": [2.5] * 60,
            }
        )
        provider._memory[f"valuation:000300:{LIXINGER_SOURCE}"] = frame
        self.assertEqual(provider._best_exact_source(INDEX_SPECS["000300"]), LIXINGER_SOURCE)


class DatabaseConfigTest(unittest.TestCase):
    def test_parse_mysql_url(self):
        config = parse_mysql_url(
            "mysql://etf_user:p%40ss@127.0.0.1:3307/etf_data"
        )
        self.assertEqual(config.user, "etf_user")
        self.assertEqual(config.password, "p@ss")
        self.assertEqual(config.port, 3307)
        self.assertEqual(config.database, "etf_data")


class AnalyzerSourceTest(unittest.TestCase):
    def test_official_pe_is_used_for_absolute_valuation(self):
        dates = pd.date_range("2020-01-01", periods=100, freq="30D")
        valuation = pd.DataFrame({"date": dates, "value": [20.0] * 100})
        prices = pd.DataFrame(
            {
                "date": pd.date_range("2025-01-01", periods=300, freq="B"),
                "close": [100.0 + i / 10 for i in range(300)],
            }
        )

        class FakeProvider:
            def official_snapshot(self, spec):
                return OfficialSnapshot(
                    date="2026-08-28", pe=10.0, dividend_yield=2.0,
                    source="csindex_official",
                )

            def pe_history(self, spec):
                return valuation

            def pb_history(self, spec):
                return None

            def price_history(self, spec):
                return prices

            def using_exact(self, code):
                return False

            def history_source(self, code):
                return "legulegu_market_proxy"

            def history_quality(self, code):
                return "proxy"

        result = analyze_index(
            INDEX_SPECS["000001"], FakeProvider(), history_years=10,
            risk_free_rate=2.0,
        )
        self.assertEqual(result.analysis_pe, 10.0)
        self.assertEqual(result.earnings_yield, 10.0)
        self.assertEqual(result.analysis_pe_source, "csindex_official")
        self.assertIn("当前官方", result.confidence)


class ReportTest(unittest.TestCase):
    def test_detail_sheet_contains_separate_current_and_history_sources(self):
        metric = MetricReference(
            current=13.0, percentile=55.0, q20=10.0, median=12.0, q80=15.0,
            sample_count=120, start_date="2016-01-01", end_date="2026-01-01",
        )
        move = MarketMove(
            latest_date="2026-08-28", close=4000.0, return_1d=0.01,
            return_5d=-0.02, return_20d=0.03, drawdown_60d=-0.05,
            drawdown_250d=-0.08, volatility_20d=0.20,
        )
        result = ValuationResult(
            code="000300", name="沪深300", etf_code="510300", status="合理",
            confidence="较高（历史指数级，当前官方）", pe=metric, pb=None,
            official_pe=14.0, official_dividend_yield=2.5,
            official_date="2026-08-28", analysis_pe=14.0,
            analysis_pe_source="csindex_official",
            history_source="legulegu_index", history_quality="index",
            earnings_yield=100 / 14, risk_free_rate=2.0,
            earnings_yield_spread=100 / 14 - 2.0, composite_percentile=55.0,
            market=move, base_multiplier=1.0, dip_bonus=0.1,
            suggested_multiplier=1.1, reason="测试", note="测试来源拆分",
            target_weight=1.0, base_amount=1000.0, suggested_amount=1100.0,
        )
        workbook = Workbook()
        sheet = workbook.active
        _build_detail(sheet, [result])
        headers = [sheet.cell(3, col).value for col in range(1, sheet.max_column + 1)]
        self.assertEqual(sheet.max_column, 32)
        self.assertIn("当前PE来源", headers)
        self.assertIn("历史PE来源", headers)
        self.assertEqual(sheet.cell(4, 17).value, "csindex_official")
        self.assertEqual(sheet.cell(4, 18).value, "legulegu_index")


if __name__ == "__main__":
    unittest.main()
