from __future__ import annotations

from datetime import datetime
from pathlib import Path

from openpyxl import Workbook
from openpyxl.formatting.rule import FormulaRule
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
from openpyxl.utils import get_column_letter

from think.etf.models import ValuationResult
from think.etf.strategy import decision_text


NAVY = "17365D"
BLUE = "2F75B5"
LIGHT_BLUE = "D9EAF7"
GREEN = "E2F0D9"
YELLOW = "FFF2CC"
ORANGE = "FCE4D6"
RED = "F4CCCC"
GRAY = "E7E6E6"
WHITE = "FFFFFF"
BLACK = "000000"
THIN_GRAY = Side(style="thin", color="D9E1F2")


def export_excel_report(
    results: list[ValuationResult],
    output_path: str | Path,
    risk_free_rate: float,
    risk_free_date: str,
    warnings: list[str],
    failures: list[str],
    database_label: str = "未提供",
) -> Path:
    path = Path(output_path)
    path.parent.mkdir(parents=True, exist_ok=True)
    workbook = Workbook()
    summary = workbook.active
    summary.title = "本周结论"
    detail = workbook.create_sheet("估值明细")
    rules = workbook.create_sheet("指标与规则")
    sources = workbook.create_sheet("数据口径")
    checks = workbook.create_sheet("检查")

    _build_summary(summary, results, risk_free_rate, risk_free_date, database_label)
    _build_detail(detail, results)
    _build_rules(rules)
    _build_sources(sources, results, warnings, failures)
    _build_checks(checks, results, failures, database_label)

    for sheet in workbook.worksheets:
        sheet.sheet_view.showGridLines = False
        sheet.sheet_properties.pageSetUpPr.fitToPage = True
        sheet.page_setup.fitToWidth = 1
        sheet.page_setup.fitToHeight = 0
    workbook.save(path)
    return path.resolve()


def default_report_path(output_dir: str | Path) -> Path:
    timestamp = datetime.now().strftime("%Y%m%d_%H%M%S")
    return Path(output_dir) / f"ETF估值定投报告_{timestamp}.xlsx"


def direct_conclusion(item: ValuationResult) -> str:
    base = decision_text(item)
    return f"{base}，本周倍率 {item.suggested_multiplier:.2f}。{item.reason}。"


def _build_summary(
    sheet, results, risk_free_rate: float, risk_free_date: str, database_label: str
) -> None:
    sheet.merge_cells("A1:K1")
    sheet["A1"] = "ETF每周定投结论"
    sheet["A1"].font = Font(name="Microsoft YaHei", size=18, bold=True, color=WHITE)
    sheet["A1"].fill = PatternFill("solid", fgColor=NAVY)
    sheet["A1"].alignment = Alignment(horizontal="left", vertical="center")
    sheet.row_dimensions[1].height = 34
    sheet.merge_cells("A2:K2")
    sheet["A2"] = (
        f"数据截至各指数最新交易日｜10年期国债收益率 {risk_free_rate:.2f}%（{risk_free_date}）｜"
        f"数据库 {database_label}｜结论用于决定本周定投金额，不代表短期涨跌预测"
    )
    sheet["A2"].fill = PatternFill("solid", fgColor=LIGHT_BLUE)
    sheet["A2"].alignment = Alignment(wrap_text=True, vertical="center")
    sheet.row_dimensions[2].height = 30

    headers = [
        "指数", "指数代码", "ETF代码", "估值状态", "是否定投", "投资倍率",
        "基础金额(元)", "建议金额(元)", "大跌加仓", "直接结论", "数据可信度",
    ]
    sheet.append([])
    sheet.append(headers)
    for item in results:
        sheet.append(
            [
                item.name,
                int(item.code),
                int(item.etf_code),
                item.status,
                decision_text(item),
                item.suggested_multiplier,
                item.base_amount,
                item.suggested_amount,
                "是" if item.dip_bonus > 0 else "否",
                direct_conclusion(item),
                item.confidence,
            ]
        )
    end_row = 4 + len(results)
    _style_header(sheet, 4, len(headers))
    sheet.freeze_panes = "A5"
    sheet.auto_filter.ref = f"A4:K{end_row}"
    sheet.column_dimensions["A"].width = 14
    sheet.column_dimensions["B"].width = 12
    sheet.column_dimensions["C"].width = 12
    sheet.column_dimensions["D"].width = 12
    sheet.column_dimensions["E"].width = 22
    sheet.column_dimensions["F"].width = 11
    sheet.column_dimensions["G"].width = 14
    sheet.column_dimensions["H"].width = 14
    sheet.column_dimensions["I"].width = 11
    sheet.column_dimensions["J"].width = 52
    sheet.column_dimensions["K"].width = 30
    for row in range(5, end_row + 1):
        sheet.row_dimensions[row].height = 38
        for col in range(1, 12):
            cell = sheet.cell(row, col)
            cell.alignment = Alignment(
                horizontal="left" if col in {1, 5, 10, 11} else "center",
                vertical="center",
                wrap_text=True,
            )
            cell.border = Border(bottom=THIN_GRAY)
        sheet.cell(row, 6).number_format = "0.00x"
        sheet.cell(row, 2).number_format = "000000"
        sheet.cell(row, 3).number_format = "000000"
        sheet.cell(row, 7).number_format = "#,##0.00"
        sheet.cell(row, 8).number_format = "#,##0.00"
        _fill_status(sheet.cell(row, 4))
    sheet.conditional_formatting.add(
        f"F5:F{end_row}",
        FormulaRule(formula=["F5>1"], fill=PatternFill("solid", fgColor=GREEN)),
    )
    sheet.conditional_formatting.add(
        f"F5:F{end_row}",
        FormulaRule(formula=["F5<1"], fill=PatternFill("solid", fgColor=ORANGE)),
    )

    note_row = end_row + 3
    sheet.merge_cells(start_row=note_row, start_column=1, end_row=note_row, end_column=11)
    sheet.cell(note_row, 1).value = "执行纪律"
    _style_section(sheet.cell(note_row, 1))
    notes = [
        "每周固定一天运行并执行一次；普通波动无需每天查看。",
        "只有出现单日≥7%、近5日≥5%或距60日高点回撤达到规则档位时，才临时重跑一次。",
        "临时重跑只有在估值为合理或更低时才允许额外加仓；偏高/明显高估仍不加仓。",
        "额外投入来自预留现金，不借贷、不使用杠杆；同一周最多执行一次临时加仓。",
    ]
    for offset, text in enumerate(notes, start=1):
        sheet.merge_cells(
            start_row=note_row + offset,
            start_column=1,
            end_row=note_row + offset,
            end_column=11,
        )
        cell = sheet.cell(note_row + offset, 1)
        cell.value = f"• {text}"
        cell.alignment = Alignment(wrap_text=True, vertical="center")
        sheet.row_dimensions[note_row + offset].height = 24


def _build_detail(sheet, results) -> None:
    sheet.merge_cells("A1:AF1")
    sheet["A1"] = "估值与市场指标明细"
    _style_title(sheet["A1"])
    sheet.row_dimensions[1].height = 32
    headers = [
        "指数代码", "指数", "ETF", "状态", "可信度", "数据日期", "PE", "精确/官方PE快照",
        "PE分位", "PE20%", "PE中位", "PE80%", "PB", "PB分位", "股息率", "评分采用PE", "当前PE来源", "历史PE来源", "盈利收益率",
        "10年国债", "股债差", "综合评分", "近1日", "近5日", "近20日", "60日回撤",
        "250日回撤", "基础倍率", "大跌加仓", "建议倍率", "原因", "口径说明",
    ]
    sheet.append([])
    sheet.append(headers)
    for item in results:
        sheet.append(
            [
                int(item.code), item.name, int(item.etf_code), item.status, item.confidence,
                item.market.latest_date, item.pe.current, item.official_pe,
                item.pe.percentile / 100, item.pe.q20, item.pe.median, item.pe.q80,
                item.pb.current if item.pb else None,
                item.pb.percentile / 100 if item.pb else None,
                item.official_dividend_yield / 100 if item.official_dividend_yield else None,
                item.analysis_pe, item.analysis_pe_source, item.history_source,
                item.earnings_yield / 100, item.risk_free_rate / 100,
                item.earnings_yield_spread / 100, item.composite_percentile / 100,
                item.market.return_1d, item.market.return_5d, item.market.return_20d,
                item.market.drawdown_60d, item.market.drawdown_250d,
                item.base_multiplier, item.dip_bonus, item.suggested_multiplier,
                item.reason, item.note,
            ]
        )
    end_row = 3 + len(results)
    _style_header(sheet, 3, len(headers))
    sheet.freeze_panes = "G4"
    sheet.auto_filter.ref = f"A3:AF{end_row}"
    for row in range(4, end_row + 1):
        for col in range(1, 33):
            cell = sheet.cell(row, col)
            cell.border = Border(bottom=THIN_GRAY)
            cell.alignment = Alignment(vertical="center", wrap_text=col in {17, 18, 31, 32})
        for col in (9, 14, 15, 19, 20, 21, 22, 23, 24, 25, 26, 27):
            sheet.cell(row, col).number_format = "0.00%"
        for col in (7, 8, 10, 11, 12, 13, 16):
            sheet.cell(row, col).number_format = "0.00"
        sheet.cell(row, 1).number_format = "000000"
        sheet.cell(row, 3).number_format = "000000"
        for col in (28, 29, 30):
            sheet.cell(row, col).number_format = "0.00x"
        _fill_status(sheet.cell(row, 4))
    widths = {
        "A": 12, "B": 14, "C": 11, "D": 12, "E": 18, "F": 13,
        "G": 10, "H": 17, "I": 11, "J": 10, "K": 10, "L": 10,
        "M": 10, "N": 11, "O": 11, "P": 13, "Q": 11, "R": 11,
        "S": 11, "T": 10, "U": 10, "V": 10, "W": 11, "X": 12,
        "Y": 10, "Z": 11, "AA": 12, "AB": 11, "AC": 11, "AD": 11,
        "AE": 34, "AF": 58,
    }
    for column, width in widths.items():
        sheet.column_dimensions[column].width = width
    for row in range(4, end_row + 1):
        sheet.row_dimensions[row].height = 45


def _build_rules(sheet) -> None:
    sheet.merge_cells("A1:F1")
    sheet["A1"] = "指标含义与判断规则"
    _style_title(sheet["A1"])
    sheet.row_dimensions[1].height = 32
    rows = [
        ["指标", "含义", "低估参考", "高估参考", "注意事项", "程序用途"],
        ["PE", "指数市值÷指数盈利", "低于历史20%分位", "高于历史80%分位", "亏损及周期顶部会失真", "相对估值主指标"],
        ["PB", "指数市值÷指数净资产", "低于历史20%分位", "高于历史80%分位", "应结合ROE和资产质量", "PE的辅助验证"],
        ["盈利收益率", "1÷PE", "越高通常越便宜", "越低通常越贵", "不是未来保证收益率", "与国债收益率比较"],
        ["股债差", "盈利收益率－10年国债收益率", "差值越大补偿越高", "低于1%补偿很弱", "成长指数还需考虑盈利增长", "绝对估值评分"],
        ["股息率", "年度现金分红÷指数市值", "高且可持续更有吸引力", "过低或分红不可持续", "ETF是否现金分红仍看基金合同", "现金回报参考"],
        ["60日回撤", "当前收盘距60日最高收盘的跌幅", "回撤扩大可能提高投入", "不代表一定低估", "偏高估值时不触发加仓", "定投节奏"],
    ]
    for row in rows:
        sheet.append(row)
    _style_header(sheet, 2, 6)
    sheet.freeze_panes = "A3"
    for row in range(3, 3 + len(rows) - 1):
        sheet.row_dimensions[row].height = 42
        for col in range(1, 7):
            sheet.cell(row, col).alignment = Alignment(wrap_text=True, vertical="center")
            sheet.cell(row, col).border = Border(bottom=THIN_GRAY)
    for col, width in enumerate([17, 30, 24, 24, 40, 24], start=1):
        sheet.column_dimensions[get_column_letter(col)].width = width

    start = 11
    sheet.merge_cells(start_row=start, start_column=1, end_row=start, end_column=6)
    sheet.cell(start, 1).value = "10%、6.4%和6.8%到底指什么"
    _style_section(sheet.cell(start, 1))
    explanations = [
        "10%：指盈利收益率，不是股息率。对应PE=10倍，属于严格低估经验线。",
        "你上文原始规则写的是6.4%：它同样指盈利收益率，对应PE约15.63倍；来源是旧资料中债券基金长期平均收益的经验值。",
        "如果你说的6.8%来自“无风险利率约3.4%的两倍”，它仍然是盈利收益率门槛，对应PE约14.71倍。",
        "这些固定数字产生于旧利率环境，程序没有机械使用；当前使用盈利收益率与实时10年国债收益率的差值。",
    ]
    for offset, text in enumerate(explanations, start=1):
        sheet.merge_cells(
            start_row=start + offset, start_column=1, end_row=start + offset, end_column=6
        )
        cell = sheet.cell(start + offset, 1)
        cell.value = f"• {text}"
        cell.alignment = Alignment(wrap_text=True, vertical="center")
        sheet.row_dimensions[start + offset].height = 30


def _build_sources(sheet, results, warnings, failures) -> None:
    sheet.merge_cells("A1:F1")
    sheet["A1"] = "数据来源与精确数据替换"
    _style_title(sheet["A1"])
    sheet.row_dimensions[1].height = 32
    headers = ["指数", "代码", "当前口径", "MySQL存储", "如何补齐精确历史", "备注"]
    sheet.append([])
    sheet.append(headers)
    for item in results:
        storage = f"etf_index_valuation.index_code={item.code}"
        if item.history_source == "lixinger_index_mcw":
            method = (
                "理杏仁开放平台已自动同步指数市值加权PE/PB/股息率；"
                "后续运行自动增量更新MySQL"
            )
        else:
            method = (
                "从指数公司或持牌数据库导出到任意外部路径，再用"
                "--import-valuation-csv导入MySQL；CSV字段为date,pe,pb,dividend_yield"
            )
        sheet.append([item.name, int(item.code), item.confidence, storage, method, item.note])
    end_row = 3 + len(results)
    _style_header(sheet, 3, 6)
    sheet.freeze_panes = "A4"
    for row in range(4, end_row + 1):
        sheet.row_dimensions[row].height = 58
        for col in range(1, 7):
            sheet.cell(row, col).alignment = Alignment(wrap_text=True, vertical="center")
            sheet.cell(row, col).border = Border(bottom=THIN_GRAY)
        sheet.cell(row, 2).number_format = "000000"
    widths = [16, 12, 20, 38, 60, 70]
    for col, width in enumerate(widths, start=1):
        sheet.column_dimensions[get_column_letter(col)].width = width

    source_row = end_row + 3
    sheet.merge_cells(start_row=source_row, start_column=1, end_row=source_row, end_column=6)
    sheet.cell(source_row, 1).value = "参考网址"
    _style_section(sheet.cell(source_row, 1))
    urls = [
        "https://www.lixinger.com/api/open-api/html-doc/cn/index/fundamental（理杏仁指数基本面API）",
        "https://www.csindex.com.cn/（中证指数：上证、沪深300、红利、科创100）",
        "https://www.cnindex.com.cn/（国证指数：创业板指）",
        "https://yield.chinabond.com.cn/（中债国债收益率曲线）",
        "https://etf.sse.com.cn/（上交所ETF信息）",
        "https://www.szse.cn/（深交所ETF信息）",
    ]
    for offset, url in enumerate(urls, start=1):
        sheet.merge_cells(
            start_row=source_row + offset, start_column=1,
            end_row=source_row + offset, end_column=6
        )
        sheet.cell(source_row + offset, 1).value = url
    if warnings or failures:
        warning_row = source_row + len(urls) + 2
        sheet.merge_cells(start_row=warning_row, start_column=1, end_row=warning_row, end_column=6)
        sheet.cell(warning_row, 1).value = "运行警告"
        _style_section(sheet.cell(warning_row, 1))
        for offset, text in enumerate(warnings + failures, start=1):
            sheet.merge_cells(
                start_row=warning_row + offset, start_column=1,
                end_row=warning_row + offset, end_column=6
            )
            sheet.cell(warning_row + offset, 1).value = f"• {text}"


def _build_checks(sheet, results, failures, database_label: str) -> None:
    sheet.merge_cells("A1:F1")
    sheet["A1"] = "数据完整性检查"
    _style_title(sheet["A1"])
    sheet.row_dimensions[1].height = 32
    headers = ["检查项目", "实际", "期望", "差异/风险", "状态", "处理建议"]
    sheet.append([])
    sheet.append(headers)
    rows = [
        ["数据库模式", database_label, "MySQL已连接", "未连接时不保存历史",
         "OK" if database_label.startswith("MySQL") else "警告",
         "日常运行请设置ETF_DATABASE_URL"],
        ["指数数量", len(results), 5, len(results) - 5, "OK" if len(results) == 5 else "警告", "检查失败列表"],
        ["运行失败数量", len(failures), 0, len(failures), "OK" if not failures else "警告", "查看数据口径页"],
        ["精确历史估值数量", sum("精确" in x.confidence for x in results), 5,
         sum("精确" in x.confidence for x in results) - 5,
         "OK" if all("精确" in x.confidence for x in results) else "提示",
         "将指数级历史估值CSV导入MySQL"],
        ["缺少PB的指数数量", sum(x.pb is None for x in results), 0, sum(x.pb is None for x in results),
         "OK" if all(x.pb is not None for x in results) else "提示", "PB缺失时程序只用PE相对分位"],
    ]
    for row in rows:
        sheet.append(row)
    _style_header(sheet, 3, 6)
    for row in range(4, 4 + len(rows)):
        for col in range(1, 7):
            sheet.cell(row, col).alignment = Alignment(wrap_text=True, vertical="center")
            sheet.cell(row, col).border = Border(bottom=THIN_GRAY)
        status = sheet.cell(row, 5)
        status.fill = PatternFill("solid", fgColor=GREEN if status.value == "OK" else YELLOW)
    for col, width in enumerate([28, 14, 14, 18, 12, 45], start=1):
        sheet.column_dimensions[get_column_letter(col)].width = width


def _style_title(cell) -> None:
    cell.font = Font(name="Microsoft YaHei", size=17, bold=True, color=WHITE)
    cell.fill = PatternFill("solid", fgColor=NAVY)
    cell.alignment = Alignment(horizontal="left", vertical="center")


def _style_section(cell) -> None:
    cell.font = Font(name="Microsoft YaHei", bold=True, color=WHITE)
    cell.fill = PatternFill("solid", fgColor=BLUE)
    cell.alignment = Alignment(vertical="center")


def _style_header(sheet, row: int, columns: int) -> None:
    for col in range(1, columns + 1):
        cell = sheet.cell(row, col)
        cell.font = Font(name="Microsoft YaHei", bold=True, color=WHITE)
        cell.fill = PatternFill("solid", fgColor=BLUE)
        cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
    sheet.row_dimensions[row].height = 32


def _fill_status(cell) -> None:
    colors = {
        "明显低估": GREEN,
        "偏低": "C6E0B4",
        "合理": YELLOW,
        "偏高": ORANGE,
        "明显高估": RED,
    }
    cell.fill = PatternFill("solid", fgColor=colors.get(str(cell.value), GRAY))
    cell.font = Font(bold=True, color=BLACK)
